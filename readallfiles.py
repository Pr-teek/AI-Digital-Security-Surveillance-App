from PyPDF2 import PdfReader
import csv
import docx
import os
import openpyxl

#
import imaplib
import email
import os
import joblib
from pdftotext import check_file_extension, is_password_protected, decrypt_pdf

def read_file(file_path):
    try:
        if file_path.endswith(".pdf"):
            reader = PdfReader(file_path)
            page = reader.pages[0]
            text = page.extract_text()
        elif file_path.endswith(".csv"):
            with open(file_path, "r") as file:
                text = file.read()
        elif file_path.endswith(".docx"):
            doc = docx.Document(file_path)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        elif file_path.endswith(".xlsx"):
            wb = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows():
                    row_text = "\t".join([str(cell.value) for cell in row])
                    text += row_text + "\n"
        elif file_path.endswith(".txt"):
            with open(file_path, "r") as file:
                text = file.read()
                print(text)
        
        else:
            raise ValueError("Unsupported file format")
        
        return text
    except Exception as e:
        print("Error:", e)
        return None

# # Example usage
# file_path ="C:/General/CSE Core/Hackathons/NH-IEEE/documents/" + "randomtext.txt" # Replace with the actual file path
# content = read_file(file_path)
# print("File content:\n", content)

# if not os.path.isfile(file_path) :
#     fp = open(file_path, 'wb')
#     fp.close()

